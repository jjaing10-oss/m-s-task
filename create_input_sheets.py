#!/usr/bin/env python3
"""
27E 시뮬레이션 입력값 취합용 엑셀 생성
- 통신사업부용 (telecom_input_27E.xlsx)
- 유통플랫폼용 (distribution_input_27E.xlsx)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 스타일 정의 ──
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
SUB_FILL    = PatternFill("solid", fgColor="D6E4F0")
SUB_FONT    = Font(name="맑은 고딕", size=10, bold=True, color="1F4E79")
INPUT_FILL  = PatternFill("solid", fgColor="FFF2CC")  # 연노랑 = 입력칸
LABEL_FONT  = Font(name="맑은 고딕", size=10)
TITLE_FONT  = Font(name="맑은 고딕", size=14, bold=True, color="1F4E79")
NOTE_FONT   = Font(name="맑은 고딕", size=9, italic=True, color="808080")
REF_FILL    = PatternFill("solid", fgColor="E2EFDA")  # 연초록 = 26E 참고값
REF_FONT    = Font(name="맑은 고딕", size=10, color="548235")
thin_border = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7'),
)
MONTHS = ["1월","2월","3월","4월","5월","6월","7월","8월","9월","10월","11월","12월","연간합계"]

def style_header_row(ws, row, max_col):
    for c in range(1, max_col+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

def style_sub_row(ws, row, max_col):
    for c in range(1, max_col+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = SUB_FILL
        cell.font = SUB_FONT
        cell.border = thin_border

def add_input_row(ws, row, col_start, col_end, is_ref=False):
    for c in range(col_start, col_end+1):
        cell = ws.cell(row=row, column=c)
        if is_ref:
            cell.fill = REF_FILL
            cell.font = REF_FONT
        else:
            cell.fill = INPUT_FILL
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='right')
        cell.number_format = '#,##0'

def add_section(ws, start_row, section_title, items, ref_values=None, unit="백만원"):
    """
    items: list of (label, note_or_None)
    ref_values: dict of {label: [m1..m12]} for 26E reference
    Returns next available row
    """
    r = start_row
    # Section header
    ws.cell(row=r, column=1, value=section_title)
    style_sub_row(ws, r, 15)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=15)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal='left', vertical='center')
    r += 1

    for label, note in items:
        # 26E reference row (if available)
        if ref_values and label in ref_values:
            ws.cell(row=r, column=1, value=f"  {label} (26E 참고)")
            ws.cell(row=r, column=1).font = REF_FONT
            ws.cell(row=r, column=1).border = thin_border
            vals = ref_values[label]
            for i, v in enumerate(vals):
                ws.cell(row=r, column=3+i, value=v)
            # annual sum
            ws.cell(row=r, column=15, value=sum(vals))
            add_input_row(ws, r, 3, 15, is_ref=True)
            r += 1

        # 27E input row
        ws.cell(row=r, column=1, value=f"  {label}")
        ws.cell(row=r, column=1).font = LABEL_FONT
        ws.cell(row=r, column=1).border = thin_border
        if note:
            ws.cell(row=r, column=2, value=note)
            ws.cell(row=r, column=2).font = NOTE_FONT
            ws.cell(row=r, column=2).border = thin_border
        # Yellow input cells for months + annual
        add_input_row(ws, r, 3, 15)
        # Annual sum formula
        ws.cell(row=r, column=15).value = f"=SUM({get_column_letter(3)}{r}:{get_column_letter(14)}{r})"
        r += 1

    return r


def create_telecom_sheet():
    """통신사업부 입력 시트"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "통신_27E_입력"

    # Column widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 20
    for c in range(3, 16):
        ws.column_dimensions[get_column_letter(c)].width = 12

    # ── Title ──
    ws.cell(row=1, column=1, value="KT M&S 통신사업부 — 27E 시뮬레이션 입력값")
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.merge_cells('A1:O1')
    ws.row_dimensions[1].height = 30

    ws.cell(row=2, column=1, value="※ 노란칸에 월별 값을 입력해 주세요. 초록칸은 26E 참고값입니다.")
    ws.cell(row=2, column=1).font = NOTE_FONT
    ws.merge_cells('A2:O2')

    # ── Header row ──
    r = 4
    ws.cell(row=r, column=1, value="항목")
    ws.cell(row=r, column=2, value="단위/비고")
    for i, m in enumerate(MONTHS):
        ws.cell(row=r, column=3+i, value=m)
    style_header_row(ws, r, 15)
    r += 1

    # ── 1. 무선 CAPA ──
    ref_capa = {
        "무선 총 CAPA": [55845,55845,55845,55845,55845,55845,55845,55845,55845,55845,55845,55845],
        "  도매": [28419,20844,25288,24174,24539,24152,24500,22244,22244,22244,22244,20244],
        "  소매": [17400,17400,17200,17200,17200,17200,17200,17200,17200,17200,17200,17200],
    }
    r = add_section(ws, r, "1. 무선 CAPA (가입자 순증)", [
        ("무선 총 CAPA", "건"),
        ("  도매", "건"),
        ("  소매", "건"),
        ("  소상공인", "건"),
        ("  디지털", "건"),
        ("  기업", "건"),
        ("  Biz", "건"),
        ("  IoT", "건"),
    ], ref_values={
        "무선 총 CAPA": [55845]*12,
    })

    r += 1  # blank row

    # ── 2. 유선 ──
    r = add_section(ws, r, "2. 유선", [
        ("유선 순신규 합계", "건"),
        ("  인터넷", "건"),
        ("  IPTV", "건"),
        ("  전화", "건"),
        ("유선 정책수수료 (총액)", "백만원 (26E: 11,369)"),
    ])

    r += 1

    # ── 3. 대당 단가 (무선) ──
    r = add_section(ws, r, "3. 무선 대당 단가 (단가 변동 있을 시)", [
        ("대당 상품매출 (PU)", "원/건 (26E: 708,268)"),
        ("대당 수수료매출 (CU)", "원/건 (26E: 760,673)"),
        ("무선 대당 정책수수료", "원/건 (26E: 247,338)"),
        ("대당 관리수수료", "원/매장"),
    ])

    r += 1

    # ── 4. 매장수 ──
    r = add_section(ws, r, "4. 매장수", [
        ("소매 매장수", "개"),
        ("도매 매장수", "개"),
    ], ref_values={
        "소매 매장수": [251,251,251,252,254,256,258,260,262,264,266,270],
        "도매 매장수": [42,41,41,40,40,39,39,38,38,36,34,34],
    })

    r += 1

    # ── 5. 인력수 ──
    r = add_section(ws, r, "5. 인력 (HC)", [
        ("정규직 인원", "명"),
        ("영업직 인원", "명"),
        ("도급/계약직 인원", "명"),
        ("임금인상률 (연간)", "% (예: 3.5)"),
    ])

    r += 1

    # ── 6. 원가/비용 가정 ──
    r_start = r
    ws.cell(row=r, column=1, value="6. 비용 가정 (연간)")
    style_sub_row(ws, r, 15)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=15)
    r += 1

    annual_params = [
        ("통신 원가율 개선 목표", "% (예: 0.5 = 0.5%p 개선)"),
        ("CAPEX 투자계획", "백만원"),
        ("감가상각 내용연수", "년"),
        ("임차료 증감률", "% (예: 2.5)"),
        ("마케팅비 증감률", "% (예: 2.0)"),
        ("판촉비 증감률", "% (예: 3.0)"),
    ]
    for label, note in annual_params:
        ws.cell(row=r, column=1, value=f"  {label}")
        ws.cell(row=r, column=1).font = LABEL_FONT
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=2, value=note)
        ws.cell(row=r, column=2).font = NOTE_FONT
        ws.cell(row=r, column=2).border = thin_border
        ws.cell(row=r, column=3).fill = INPUT_FILL
        ws.cell(row=r, column=3).border = thin_border
        r += 1

    # Freeze panes
    ws.freeze_panes = 'C5'

    wb.save("/home/user/m-s-task/telecom_input_27E.xlsx")
    print("✅ telecom_input_27E.xlsx 생성 완료")


def create_distribution_sheet():
    """유통플랫폼 입력 시트"""
    wb = openpyxl.Workbook()

    channels = [
        ("goodbuy", "굿바이", "상품", "중고폰 매입/판매"),
        ("siyon", "시연폰", "상품", "시연폰 매입/재판매"),
        ("otherbiz", "기타사업", "상품", "중고폰전문매장 등"),
        ("coconut", "코코넛", "서비스", "보험/보상 서비스"),
        ("aslog", "AS물류대행", "서비스", "AS 물류 대행"),
        ("rental", "임대폰대행", "서비스", "임대폰 대행"),
    ]

    # 26E 참고값 (DP_SECTION_26 연간 data)
    ref_26 = {
        "goodbuy": {"상품매출": 15527, "정책수수료": 2360, "기타수수료": 225,
                     "상품원가": 14010, "인건비": 416, "임차료": 28,
                     "지급수수료": 1865, "판촉비": 687, "기타운영비": 411, "영업이익": 645},
        "siyon":   {"상품매출": 36845, "정책수수료": 226, "기타수수료": 48,
                     "매각원가": 29985, "인건비": 269, "임차료": 31,
                     "지급수수료": 816, "판촉비": 126, "기타운영비": 6, "영업이익": 5862},
        "otherbiz":{"상품매출": 2700, "정책수수료": 320,
                     "상품원가": 1600, "인건비": 574, "임차료": 90,
                     "지급수수료": 448, "판촉비": 56, "기타운영비": 40, "영업이익": 212},
        "coconut": {"상품매출": 483, "정책수수료": 224, "기타수수료": 20,
                     "상품원가": 215, "인건비": 385, "감가상각": 37, "영업이익": 0},
        "aslog":   {"기타수수료": 360,
                     "인건비": 38, "임차료": 15, "지급수수료": 233,
                     "운반비": 10, "기타운영비": 10, "영업이익": 45},
        "rental":  {"기타수수료": 442,
                     "인건비": 40, "임차료": 11, "지급수수료": 70,
                     "운반비": 15, "기타운영비": 25, "영업이익": 270},
    }

    # ── 총괄 시트 ──
    ws_sum = wb.active
    ws_sum.title = "유통_총괄_27E"

    ws_sum.column_dimensions['A'].width = 28
    ws_sum.column_dimensions['B'].width = 18
    for c in range(3, 16):
        ws_sum.column_dimensions[get_column_letter(c)].width = 12

    ws_sum.cell(row=1, column=1, value="KT M&S 유통플랫폼 — 27E 시뮬레이션 입력값 총괄")
    ws_sum.cell(row=1, column=1).font = TITLE_FONT
    ws_sum.merge_cells('A1:O1')
    ws_sum.row_dimensions[1].height = 30

    ws_sum.cell(row=2, column=1, value="※ 각 채널 시트에 월별 상세를 입력하시면 이 시트에 자동 합산됩니다.")
    ws_sum.cell(row=2, column=1).font = NOTE_FONT
    ws_sum.merge_cells('A2:O2')

    # Header
    r = 4
    ws_sum.cell(row=r, column=1, value="채널")
    ws_sum.cell(row=r, column=2, value="26E 연간 (참고)")
    ws_sum.cell(row=r, column=3, value="27E 물량 (연간)")
    ws_sum.cell(row=r, column=4, value="27E 매장수")
    ws_sum.cell(row=r, column=5, value="27E 인력(HC)")
    ws_sum.cell(row=r, column=6, value="27E 매출 (예상)")
    ws_sum.cell(row=r, column=7, value="비고")
    style_header_row(ws_sum, r, 7)
    r += 1

    summary_data = [
        ("굿바이", "18,112", "", "", "", "", "중고폰 매입물량 기준"),
        ("시연폰", "37,119", "", "", "", "", "시연폰 출하량 기준"),
        ("기타사업", "3,020", "", "", "", "", "중고폰전문매장 물량"),
        ("코코넛", "727", "", "", "", "", "보험건수 or 고정매출"),
        ("AS물류대행", "360", "", "", "", "", "대행 계약건수"),
        ("임대폰대행", "442", "", "", "", "", "임대폰 계약건수"),
    ]
    for ch_name, rev26, *rest in summary_data:
        ws_sum.cell(row=r, column=1, value=ch_name)
        ws_sum.cell(row=r, column=1).font = LABEL_FONT
        ws_sum.cell(row=r, column=1).border = thin_border
        ws_sum.cell(row=r, column=2, value=rev26)
        ws_sum.cell(row=r, column=2).font = REF_FONT
        ws_sum.cell(row=r, column=2).fill = REF_FILL
        ws_sum.cell(row=r, column=2).border = thin_border
        for i, v in enumerate(rest):
            ws_sum.cell(row=r, column=3+i, value=v)
            ws_sum.cell(row=r, column=3+i).fill = INPUT_FILL
            ws_sum.cell(row=r, column=3+i).border = thin_border
        r += 1

    r += 2
    ws_sum.cell(row=r, column=1, value="※ 공통 가정")
    ws_sum.cell(row=r, column=1).font = SUB_FONT
    r += 1
    common_params = [
        ("유통 임금인상률", "% (예: 3.5)"),
        ("유통 임차료 증감률", "% (예: 2.5)"),
        ("유통 전체 성장률 (기본)", "% (예: 2.0)"),
    ]
    for label, note in common_params:
        ws_sum.cell(row=r, column=1, value=f"  {label}")
        ws_sum.cell(row=r, column=1).font = LABEL_FONT
        ws_sum.cell(row=r, column=1).border = thin_border
        ws_sum.cell(row=r, column=2, value=note)
        ws_sum.cell(row=r, column=2).font = NOTE_FONT
        ws_sum.cell(row=r, column=2).border = thin_border
        ws_sum.cell(row=r, column=3).fill = INPUT_FILL
        ws_sum.cell(row=r, column=3).border = thin_border
        r += 1

    # ── 채널별 시트 생성 ──
    for ch_key, ch_name, ch_type, ch_desc in channels:
        ws = wb.create_sheet(title=ch_name)

        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 18
        for c in range(3, 16):
            ws.column_dimensions[get_column_letter(c)].width = 12

        # Title
        ws.cell(row=1, column=1, value=f"{ch_name} ({ch_desc}) — 27E 입력")
        ws.cell(row=1, column=1).font = TITLE_FONT
        ws.merge_cells('A1:O1')
        ws.row_dimensions[1].height = 28

        type_label = "상품 채널 → 물량(판매대수) 기반 예측" if ch_type == "상품" else "서비스 채널 → 계약/고정 매출 기반"
        ws.cell(row=2, column=1, value=f"채널 유형: {type_label}")
        ws.cell(row=2, column=1).font = NOTE_FONT
        ws.merge_cells('A2:O2')

        ws.cell(row=3, column=1, value="※ 노란칸 입력, 초록칸 = 26E Plan 참고값 (백만원)")
        ws.cell(row=3, column=1).font = NOTE_FONT

        # Header
        r = 5
        ws.cell(row=r, column=1, value="항목")
        ws.cell(row=r, column=2, value="단위/비고")
        for i, m in enumerate(MONTHS):
            ws.cell(row=r, column=3+i, value=m)
        style_header_row(ws, r, 15)
        r += 1

        # ── 물량/드라이버 ──
        if ch_type == "상품":
            r = add_section(ws, r, "1. 물량 (드라이버)", [
                ("판매/출하 물량", "건/대"),
                ("대당 매출단가", "원/건"),
                ("대당 원가", "원/건"),
            ])
        else:
            r = add_section(ws, r, "1. 계약/물량", [
                ("계약건수 or 고정매출", "건 or 백만원"),
            ])

        r += 1

        # ── 인력/매장 ──
        r = add_section(ws, r, "2. 인력 & 매장", [
            ("인원수 (HC)", "명"),
            ("매장수", "개 (해당시)"),
        ])

        r += 1

        # ── 손익 항목 (26E 참고 + 27E 입력) ──
        ref = ref_26.get(ch_key, {})

        if ch_type == "상품":
            pnl_items = [
                ("상품매출", "백만원"),
                ("광고수익", "백만원"),
                ("정책수수료", "백만원"),
                ("기타수수료", "백만원"),
                ("상품원가", "백만원"),
                ("매각원가", "백만원"),
                ("인건비", "백만원"),
                ("임차료", "백만원"),
                ("지급수수료", "백만원"),
                ("운반비", "백만원"),
                ("판촉비", "백만원"),
                ("기타운영비", "백만원"),
                ("감가상각비", "백만원"),
            ]
        else:
            pnl_items = [
                ("기타수수료", "백만원"),
                ("상품매출", "백만원 (해당시)"),
                ("정책수수료", "백만원 (해당시)"),
                ("상품원가", "백만원 (해당시)"),
                ("인건비", "백만원"),
                ("임차료", "백만원"),
                ("지급수수료", "백만원"),
                ("운반비", "백만원"),
                ("판촉비", "백만원"),
                ("기타운영비", "백만원"),
                ("감가상각비", "백만원"),
            ]

        # Build ref_values for section
        ref_annual = {}
        for label, _ in pnl_items:
            if label in ref:
                ref_annual[label] = ref[label]

        # For annual ref, put it as a note on col B
        ws.cell(row=r, column=1, value="3. 손익 항목 (월별 or 연간)")
        style_sub_row(ws, r, 15)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=15)
        r += 1

        for label, unit in pnl_items:
            # 26E reference (annual only, in col B)
            ref_val = ref.get(label, "")
            ref_text = f"26E: {ref_val:,}" if isinstance(ref_val, (int,float)) and ref_val else ""

            ws.cell(row=r, column=1, value=f"  {label}")
            ws.cell(row=r, column=1).font = LABEL_FONT
            ws.cell(row=r, column=1).border = thin_border
            ws.cell(row=r, column=2, value=f"{unit} {ref_text}")
            ws.cell(row=r, column=2).font = NOTE_FONT
            ws.cell(row=r, column=2).border = thin_border
            add_input_row(ws, r, 3, 15)
            ws.cell(row=r, column=15).value = f"=SUM({get_column_letter(3)}{r}:{get_column_letter(14)}{r})"
            r += 1

        r += 1
        ws.cell(row=r, column=1, value="※ 월별 입력이 어려우면 연간합계(O열)만 입력 → 균등배분 처리됩니다.")
        ws.cell(row=r, column=1).font = NOTE_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=15)

        ws.freeze_panes = 'C6'

    wb.save("/home/user/m-s-task/distribution_input_27E.xlsx")
    print("✅ distribution_input_27E.xlsx 생성 완료")


if __name__ == "__main__":
    create_telecom_sheet()
    create_distribution_sheet()
